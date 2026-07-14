"""Immutable-original PDF ingestion and reproducible derivative generation.

The source package has one authority boundary: ``original/`` is append-only.
Everything under ``derived/`` can be deleted and rebuilt from the original PDF
and its source manifest.  This module deliberately does not mark a source as
human-verified or citation-ready; those are governance decisions, not parser
side effects.
"""

from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any, Callable

import fitz


PIPELINE_VERSION = "1.1.0"
MANIFEST_NAME = "source-manifest.json"


class SourceIntegrityError(RuntimeError):
    """Raised when a proposed write would mutate a quarantined original."""


@dataclass(frozen=True)
class OcrResult:
    status: str
    text: str
    detail: str | None = None


def sha256_file(path: str | Path) -> str:
    """Return the byte-level SHA-256 hash without loading the whole file."""

    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_json(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _read_manifest(package_dir: Path) -> dict[str, Any]:
    manifest_path = package_dir / MANIFEST_NAME
    if not manifest_path.is_file():
        raise FileNotFoundError(f"Source manifest is missing: {manifest_path}")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def _now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def ingest_pdf(
    input_pdf: str | Path,
    package_dir: str | Path,
    *,
    source_id: str,
    rights_status: str = "unknown",
    storage_mode: str = "local_only",
) -> dict[str, Any]:
    """Quarantine an original PDF exactly once and create its source manifest.

    A repeated ingestion is idempotent only when the original bytes are
    identical.  A different file at the same source package is rejected
    instead of replacing the immutable original.
    """

    input_path = Path(input_pdf).resolve()
    package = Path(package_dir).resolve()
    if input_path.suffix.lower() != ".pdf":
        raise ValueError(f"Expected a PDF, got {input_path.name}")
    if not input_path.is_file():
        raise FileNotFoundError(input_path)
    if rights_status not in {"cleared", "restricted", "prohibited", "unknown"}:
        raise ValueError(f"Unsupported rights status: {rights_status}")
    if storage_mode not in {"committed", "local_only", "link_only", "unknown"}:
        raise ValueError(f"Unsupported storage mode: {storage_mode}")

    source_hash = sha256_file(input_path)
    original_dir = package / "original"
    target = original_dir / input_path.name
    manifest_path = package / MANIFEST_NAME

    if target.exists():
        if sha256_file(target) != source_hash:
            raise SourceIntegrityError(
                f"Refusing to overwrite immutable original at {target}; create a new source package/version instead."
            )
    else:
        original_dir.mkdir(parents=True, exist_ok=True)
        # ``copyfile`` only occurs after the existence check; the original is
        # never re-opened for writing by derivative operations.
        shutil.copyfile(input_path, target)

    if manifest_path.exists():
        existing = _read_manifest(package)
        if existing["original_sha256"] != source_hash or existing["original_filename"] != input_path.name:
            raise SourceIntegrityError("Existing source manifest conflicts with quarantined original")
        return existing

    document = fitz.open(target)
    try:
        page_count = document.page_count
    finally:
        document.close()
    manifest = {
        "schema_version": "1.0.0",
        "source_id": source_id,
        "original_filename": input_path.name,
        "original_relative_path": f"original/{input_path.name}",
        "original_sha256": source_hash,
        "original_size_bytes": target.stat().st_size,
        "page_count": page_count,
        "rights_status": rights_status,
        "storage_mode": storage_mode,
        "pdf_status": "quarantine",
        "ingested_at": _now(),
        "pipeline_version": PIPELINE_VERSION,
    }
    _write_json(manifest_path, manifest)
    return manifest


def ingest_companion_artifact(
    input_artifact: str | Path,
    package_dir: str | Path,
    *,
    artifact_id: str,
    role: str = "structured_companion",
    media_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> dict[str, Any]:
    """Add one immutable companion artifact to an existing source package.

    SDTMIG uses the PDF for complete narrative guidance and a companion XLSX
    for normative tabular metadata. They share one logical source version but
    retain independent byte hashes and locators. Re-ingestion is idempotent for
    identical bytes and rejects replacement at the same artifact identity.
    """

    input_path = Path(input_artifact).resolve()
    package = Path(package_dir).resolve()
    if not input_path.is_file():
        raise FileNotFoundError(input_path)
    if not artifact_id or not role or not media_type:
        raise ValueError("artifact_id, role, and media_type must be non-empty")

    manifest = _read_manifest(package)
    artifact_hash = sha256_file(input_path)
    target = package / "original" / input_path.name
    if target.exists():
        if sha256_file(target) != artifact_hash:
            raise SourceIntegrityError(
                f"Refusing to overwrite immutable companion at {target}; "
                "create a new source package/version instead."
            )
    else:
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(input_path, target)

    primary_artifact_id = f"artifact-{manifest['source_id'].removeprefix('src-')}-pdf"
    artifacts = list(manifest.get("artifacts", []))
    if not artifacts:
        artifacts.append(
            {
                "artifact_id": primary_artifact_id,
                "role": "primary_citation",
                "media_type": "application/pdf",
                "original_filename": manifest["original_filename"],
                "original_relative_path": manifest["original_relative_path"],
                "original_sha256": manifest["original_sha256"],
                "original_size_bytes": manifest["original_size_bytes"],
                "page_count": manifest["page_count"],
            }
        )

    existing = next((item for item in artifacts if item["artifact_id"] == artifact_id), None)
    candidate = {
        "artifact_id": artifact_id,
        "role": role,
        "media_type": media_type,
        "original_filename": input_path.name,
        "original_relative_path": f"original/{input_path.name}",
        "original_sha256": artifact_hash,
        "original_size_bytes": target.stat().st_size,
    }
    if existing is not None and existing != candidate:
        raise SourceIntegrityError(f"Existing artifact metadata conflicts for {artifact_id}")
    if existing is None:
        artifacts.append(candidate)

    manifest.update(
        {
            "schema_version": "1.1.0",
            "primary_artifact_id": primary_artifact_id,
            "artifacts": sorted(artifacts, key=lambda item: item["artifact_id"]),
            "pipeline_version": PIPELINE_VERSION,
            "manifest_updated_at": _now(),
        }
    )
    _write_json(package / MANIFEST_NAME, manifest)
    return manifest


def _json_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def build_xlsx_derivative(package_dir: str | Path, *, artifact_id: str) -> dict[str, Any]:
    """Extract workbook values and sheet structure without changing the XLSX."""

    from openpyxl import load_workbook

    package = Path(package_dir).resolve()
    source = _read_manifest(package)
    artifact = next(
        (item for item in source.get("artifacts", []) if item["artifact_id"] == artifact_id),
        None,
    )
    if artifact is None:
        raise SourceIntegrityError(f"Companion artifact is not registered: {artifact_id}")
    if artifact["media_type"] != (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        raise ValueError(f"Artifact is not an XLSX workbook: {artifact_id}")

    original_path = package / artifact["original_relative_path"]
    if not original_path.is_file() or sha256_file(original_path) != artifact["original_sha256"]:
        raise SourceIntegrityError("Companion artifact is missing or its hash has changed")

    workbook = load_workbook(original_path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [
                [_json_cell_value(value) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append(
                {
                    "name": worksheet.title,
                    "state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "rows": rows,
                }
            )
        defined_names = sorted(str(name) for name in workbook.defined_names)
        epoch = workbook.epoch.isoformat()
    finally:
        workbook.close()

    extraction = {
        "schema_version": "1.0.0",
        "source_id": source["source_id"],
        "source_sha256": source["original_sha256"],
        "artifact_id": artifact_id,
        "artifact_sha256": artifact["original_sha256"],
        "epoch": epoch,
        "defined_names": defined_names,
        "sheets": sheets,
    }
    output_path = package / "derived" / "xlsx" / f"{artifact_id}.json"
    _write_json(output_path, extraction)
    outputs = {
        str(output_path.relative_to(package)).replace("\\", "/"): sha256_file(output_path)
    }
    manifest = {
        "schema_version": "1.0.0",
        "source_id": source["source_id"],
        "artifact_id": artifact_id,
        "artifact_sha256": artifact["original_sha256"],
        "pipeline_version": PIPELINE_VERSION,
        "derivation": {
            "tool": "scripts.pdf.source_pipeline.build_xlsx_derivative",
            "tool_version": PIPELINE_VERSION,
            "input_sha256": artifact["original_sha256"],
            "created_at": _now(),
        },
        "outputs": outputs,
        "output_manifest_sha256": _sha256_json(outputs),
    }
    _write_json(package / "derived" / "xlsx-manifest.json", manifest)
    return manifest


def _default_ocr(image_path: Path) -> OcrResult:
    """Attempt local OCR; unavailable engines are recorded, never ignored."""

    try:
        import pytesseract  # type: ignore[import-not-found]
        from PIL import Image
    except ImportError:
        return OcrResult("unavailable", "", "pytesseract is not installed")
    try:
        text = pytesseract.image_to_string(Image.open(image_path))
    except Exception as exc:  # tesseract binary may be absent or unusable.
        return OcrResult("unavailable", "", f"OCR engine unavailable: {exc.__class__.__name__}")
    return OcrResult("completed", text, None)


def _page_words(page: fitz.Page) -> list[dict[str, Any]]:
    words: list[dict[str, Any]] = []
    for x0, y0, x1, y1, text, block_no, line_no, word_no in page.get_text("words", sort=True):
        words.append(
            {
                "text": text,
                "bbox": [round(x0, 3), round(y0, 3), round(x1, 3), round(y1, 3)],
                "block": block_no,
                "line": line_no,
                "word": word_no,
            }
        )
    return words


def _render_page(page: fitz.Page, output_path: Path, dpi: int = 144) -> None:
    scale = dpi / 72
    pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pixmap.save(str(output_path))


def build_derived_package(
    package_dir: str | Path,
    *,
    ocr: bool = True,
    ocr_runner: Callable[[Path], OcrResult] | None = None,
    dpi: int = 144,
) -> dict[str, Any]:
    """Build text, page-coordinate, render, and crop derivatives from a PDF.

    The resulting manifest contains hashes for every derivative and enough
    parameters to rebuild it.  A scanned PDF with no usable text layer is
    explicitly ``ocr_unavailable`` when local OCR cannot run, so it never
    silently becomes citation-ready.
    """

    package = Path(package_dir).resolve()
    source = _read_manifest(package)
    original_path = package / source["original_relative_path"]
    if not original_path.is_file():
        raise SourceIntegrityError(f"Quarantined original is missing: {original_path}")
    original_hash = sha256_file(original_path)
    if original_hash != source["original_sha256"]:
        raise SourceIntegrityError("Quarantined original hash does not match source manifest")

    derived = package / "derived"
    render_dir = derived / "render"
    document = fitz.open(original_path)
    pages: list[dict[str, Any]] = []
    full_text_parts: list[str] = []
    figures: list[dict[str, Any]] = []
    try:
        for number, page in enumerate(document, start=1):
            words = _page_words(page)
            page_text = " ".join(word["text"] for word in words)
            full_text_parts.append(page_text)
            rect = page.rect
            render_path = render_dir / f"page-{number:03d}.png"
            _render_page(page, render_path, dpi=dpi)
            page_record = {
                "physical_page": number,
                "printed_page": None,
                "width_points": round(rect.width, 3),
                "height_points": round(rect.height, 3),
                "text": page_text,
                "words": words,
                "render_relative_path": str(render_path.relative_to(package)).replace("\\", "/"),
                "render_sha256": sha256_file(render_path),
            }
            pages.append(page_record)

            # A full-page crop is a stable visual-evidence fallback even when
            # the source contains vector art rather than extractable bitmaps.
            crop_path = render_path
            figures.append(
                {
                    "figure_id": f"fig-{source['source_id']}-p{number:03d}-full",
                    "physical_page": number,
                    "bbox": [0.0, 0.0, round(rect.width, 3), round(rect.height, 3)],
                    "relative_path": str(crop_path.relative_to(package)).replace("\\", "/"),
                    "sha256": sha256_file(crop_path),
                    "kind": "page_crop",
                }
            )
    finally:
        document.close()

    has_text_layer = any(page["words"] for page in pages)
    ocr_result = OcrResult("not_required", "", None)
    if not has_text_layer and ocr:
        ocr_input = render_dir / "page-001.png"
        ocr_result = (ocr_runner or _default_ocr)(ocr_input)
        if ocr_result.status not in {"completed", "unavailable", "failed"}:
            raise ValueError(f"Unexpected OCR status: {ocr_result.status}")
        if ocr_result.status == "completed":
            full_text_parts = [ocr_result.text]
    elif not has_text_layer:
        ocr_result = OcrResult("disabled", "", "OCR disabled by caller")

    text_path = derived / "text.txt"
    text_path.parent.mkdir(parents=True, exist_ok=True)
    text_path.write_text("\n\f\n".join(full_text_parts) + "\n", encoding="utf-8")
    extraction = {
        "source_id": source["source_id"],
        "source_sha256": original_hash,
        "source_type": "digital" if has_text_layer else "scanned",
        "pages": pages,
        "ocr": asdict(ocr_result),
    }
    extraction_path = derived / "extraction.json"
    figures_path = derived / "figures.json"
    _write_json(extraction_path, extraction)
    _write_json(figures_path, {"source_id": source["source_id"], "figures": figures})

    output_paths = [text_path, extraction_path, figures_path] + [
        package / page["render_relative_path"] for page in pages
    ] + [package / figure["relative_path"] for figure in figures]
    outputs = {
        str(path.relative_to(package)).replace("\\", "/"): sha256_file(path)
        for path in sorted(output_paths)
    }
    manifest = {
        "schema_version": "1.0.0",
        "source_id": source["source_id"],
        "source_sha256": original_hash,
        "pipeline_version": PIPELINE_VERSION,
        "parameters": {"dpi": dpi, "ocr": ocr},
        "parameters_sha256": _sha256_json({"dpi": dpi, "ocr": ocr}),
        "derivation": {
            "tool": "scripts.pdf.source_pipeline",
            "tool_version": PIPELINE_VERSION,
            "input_sha256": original_hash,
            "created_at": _now(),
        },
        "source_type": extraction["source_type"],
        "ocr": asdict(ocr_result),
        "outputs": outputs,
    }
    manifest["output_manifest_sha256"] = _sha256_json(outputs)
    _write_json(derived / "manifest.json", manifest)
    return manifest
