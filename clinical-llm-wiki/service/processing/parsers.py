"""Deterministic parser adapters that preserve source locators and hashes."""

from __future__ import annotations

import csv
from hashlib import sha256
from io import BytesIO, StringIO
import json
from typing import Any, Protocol
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import fitz
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field


DOCUMENT_PARSER_PROFILE_VERSION = "deterministic-clinical-parser.v1"
_WORD_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


class ParserError(RuntimeError):
    """Base deterministic parser failure."""


class ParserQualityError(ParserError):
    """Parsed output cannot meet the evidence locator contract."""


class StrictParserModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class SourceDocument(StrictParserModel):
    source_id: str
    source_version_id: str
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    media_type: str
    object_key: str
    data_boundary: str
    rights: dict[str, Any]


class ParsedFragment(StrictParserModel):
    evidence_type: str
    locator: dict[str, Any]
    content: str = Field(min_length=1)
    content_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")

    @classmethod
    def create(
        cls,
        *,
        evidence_type: str,
        locator: dict[str, Any],
        content: str,
    ) -> "ParsedFragment":
        normalized = content.strip()
        if not normalized:
            raise ValueError("parsed fragment content cannot be empty")
        return cls(
            evidence_type=evidence_type,
            locator=locator,
            content=normalized,
            content_sha256=sha256(normalized.encode("utf-8")).hexdigest(),
        )


class ParserResult(StrictParserModel):
    contract_version: str = "parser-result.v1"
    source_id: str
    source_version_id: str
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    parser_profile_version: str = DOCUMENT_PARSER_PROFILE_VERSION
    media_type: str
    branch: str
    fragments: tuple[ParsedFragment, ...]


class ParserAdapter(Protocol):
    supported_media_types: frozenset[str]

    def parse(
        self,
        *,
        source: SourceDocument,
        content: bytes,
        branch: str,
    ) -> ParserResult: ...


class TextParser:
    supported_media_types = frozenset({"text/plain", "text/markdown"})

    def parse(
        self,
        *,
        source: SourceDocument,
        content: bytes,
        branch: str,
    ) -> ParserResult:
        if branch != "text":
            raise ParserQualityError(f"text parser does not support branch {branch}")
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise ParserQualityError("text source is not valid UTF-8") from exc
        if not text.strip():
            raise ParserQualityError("text source contains no evidence")
        line_count = len(text.splitlines()) or 1
        return _result(
            source,
            branch,
            [
                ParsedFragment.create(
                    evidence_type="text",
                    locator={
                        "kind": "line_range",
                        "start_line": 1,
                        "end_line": line_count,
                    },
                    content=text,
                )
            ],
        )


class PdfParser:
    supported_media_types = frozenset({"application/pdf"})

    def parse(
        self,
        *,
        source: SourceDocument,
        content: bytes,
        branch: str,
    ) -> ParserResult:
        try:
            document = fitz.open(stream=content, filetype="pdf")
        except Exception as exc:
            raise ParserQualityError("PDF cannot be opened deterministically") from exc
        try:
            if branch == "text":
                fragments = _pdf_text(document)
                if not fragments:
                    raise ParserQualityError(
                        "PDF contains no extractable text; OCR is required and not inferred"
                    )
            elif branch == "tables":
                fragments = _pdf_tables(document)
            elif branch == "images":
                fragments = _pdf_images(document)
            else:
                raise ParserQualityError(f"PDF parser does not support branch {branch}")
        finally:
            document.close()
        return _result(source, branch, fragments)


class DocxParser:
    supported_media_types = frozenset(
        {("application/vnd.openxmlformats-officedocument.wordprocessingml.document")}
    )

    def parse(
        self,
        *,
        source: SourceDocument,
        content: bytes,
        branch: str,
    ) -> ParserResult:
        root = _word_document_root(content)
        if branch == "text":
            fragments = []
            for index, paragraph in enumerate(root.iter(f"{_WORD_NS}p"), start=1):
                text = "".join(node.text or "" for node in paragraph.iter(f"{_WORD_NS}t")).strip()
                if text:
                    fragments.append(
                        ParsedFragment.create(
                            evidence_type="text",
                            locator={"kind": "paragraph", "paragraph": index},
                            content=text,
                        )
                    )
            if not fragments:
                raise ParserQualityError("DOCX contains no extractable paragraphs")
        elif branch == "tables":
            fragments = []
            for table_index, table in enumerate(root.iter(f"{_WORD_NS}tbl"), start=1):
                rows: list[str] = []
                for row in table.iter(f"{_WORD_NS}tr"):
                    cells = [
                        " ".join(text.text or "" for text in cell.iter(f"{_WORD_NS}t")).strip()
                        for cell in row.iter(f"{_WORD_NS}tc")
                    ]
                    rows.append("\t".join(cells))
                content_text = "\n".join(row for row in rows if row.strip())
                if content_text:
                    fragments.append(
                        ParsedFragment.create(
                            evidence_type="table",
                            locator={"kind": "table", "table": table_index},
                            content=content_text,
                        )
                    )
        else:
            raise ParserQualityError(f"DOCX parser does not support branch {branch}")
        return _result(source, branch, fragments)


class XlsxParser:
    supported_media_types = frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    def parse(
        self,
        *,
        source: SourceDocument,
        content: bytes,
        branch: str,
    ) -> ParserResult:
        if branch != "tables":
            raise ParserQualityError(f"XLSX parser does not support branch {branch}")
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=False,
            )
        except Exception as exc:
            raise ParserQualityError("XLSX cannot be opened deterministically") from exc
        fragments: list[ParsedFragment] = []
        try:
            for sheet in workbook.worksheets:
                rows = [
                    ["" if value is None else str(value) for value in row]
                    for row in sheet.iter_rows(values_only=True)
                ]
                while rows and not any(value for value in rows[-1]):
                    rows.pop()
                if not rows:
                    continue
                buffer = StringIO()
                writer = csv.writer(buffer, lineterminator="\n")
                writer.writerows(rows)
                fragments.append(
                    ParsedFragment.create(
                        evidence_type="table",
                        locator={
                            "kind": "cell_range",
                            "sheet": sheet.title,
                            "range": f"A1:{sheet.cell(len(rows), max(len(row) for row in rows)).coordinate}",
                        },
                        content=buffer.getvalue(),
                    )
                )
        finally:
            workbook.close()
        if not fragments:
            raise ParserQualityError("XLSX contains no non-empty worksheet")
        return _result(source, branch, fragments)


class ParserRegistry:
    def __init__(self, adapters: list[ParserAdapter]) -> None:
        self._by_media_type: dict[str, ParserAdapter] = {}
        for adapter in adapters:
            for media_type in adapter.supported_media_types:
                if media_type in self._by_media_type:
                    raise ValueError(f"duplicate parser media type: {media_type}")
                self._by_media_type[media_type] = adapter

    @classmethod
    def default(cls) -> "ParserRegistry":
        return cls([TextParser(), PdfParser(), DocxParser(), XlsxParser()])

    def for_media_type(self, media_type: str) -> ParserAdapter:
        try:
            return self._by_media_type[media_type]
        except KeyError as exc:
            raise ParserQualityError(f"no parser registered for {media_type}") from exc


def _result(
    source: SourceDocument,
    branch: str,
    fragments: list[ParsedFragment],
) -> ParserResult:
    return ParserResult(
        source_id=source.source_id,
        source_version_id=source.source_version_id,
        source_sha256=source.source_sha256,
        media_type=source.media_type,
        branch=branch,
        fragments=tuple(fragments),
    )


def _pdf_text(document: fitz.Document) -> list[ParsedFragment]:
    fragments: list[ParsedFragment] = []
    for page_number, page in enumerate(document, start=1):
        text = page.get_text("text").strip()
        if text:
            fragments.append(
                ParsedFragment.create(
                    evidence_type="text",
                    locator={"kind": "page", "page": page_number},
                    content=text,
                )
            )
    return fragments


def _pdf_tables(document: fitz.Document) -> list[ParsedFragment]:
    fragments: list[ParsedFragment] = []
    for page_number, page in enumerate(document, start=1):
        finder = getattr(page, "find_tables", None)
        if finder is None:
            continue
        try:
            tables = finder().tables
        except Exception:
            continue
        for table_number, table in enumerate(tables, start=1):
            extracted = table.extract()
            rows = [
                "\t".join("" if value is None else str(value) for value in row) for row in extracted
            ]
            content = "\n".join(rows).strip()
            if content:
                fragments.append(
                    ParsedFragment.create(
                        evidence_type="table",
                        locator={
                            "kind": "table",
                            "page": page_number,
                            "table": table_number,
                            "bbox": [round(value, 3) for value in table.bbox],
                        },
                        content=content,
                    )
                )
    return fragments


def _pdf_images(document: fitz.Document) -> list[ParsedFragment]:
    fragments: list[ParsedFragment] = []
    for page_number, page in enumerate(document, start=1):
        for image_number, image in enumerate(page.get_images(full=True), start=1):
            xref = image[0]
            facts = {
                "xref": xref,
                "width": image[2],
                "height": image[3],
                "colorspace": image[5],
            }
            fragments.append(
                ParsedFragment.create(
                    evidence_type="image_manifest",
                    locator={
                        "kind": "image",
                        "page": page_number,
                        "image": image_number,
                    },
                    content=json.dumps(facts, sort_keys=True, separators=(",", ":")),
                )
            )
    return fragments


def _word_document_root(content: bytes) -> ElementTree.Element:
    try:
        with ZipFile(BytesIO(content)) as archive:
            document_xml = archive.read("word/document.xml")
    except (BadZipFile, KeyError) as exc:
        raise ParserQualityError("DOCX package has no word/document.xml") from exc
    try:
        return ElementTree.fromstring(document_xml)
    except ElementTree.ParseError as exc:
        raise ParserQualityError("DOCX document XML is invalid") from exc


__all__ = [
    "DOCUMENT_PARSER_PROFILE_VERSION",
    "ParsedFragment",
    "ParserAdapter",
    "ParserError",
    "ParserQualityError",
    "ParserRegistry",
    "ParserResult",
    "SourceDocument",
]
