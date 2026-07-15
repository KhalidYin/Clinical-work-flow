"""P6-P2-A gates for the Wiki-internal source structure-map contract."""

from __future__ import annotations

from copy import deepcopy
import json
from pathlib import Path

from jsonschema import Draft202012Validator
import pytest

from scripts.pdf.create_synthetic_fixtures import create_structure_fixtures
from scripts.pdf.structure_map_contract import (
    P1_LOCATOR_FIELDS,
    StructureMapContractError,
    canonical_structure_map_bytes,
    project_locator_to_p1,
    stable_locator_id,
    stable_page_id,
    stable_reference_id,
    stable_structure_map_id,
    stable_unit_id,
    structure_map_sha256,
    validate_structure_map,
)


ROOT = Path(__file__).resolve().parents[1]
FIXTURE = ROOT / "tests" / "fixtures" / "knowledge" / "source-structure-map-positive.json"


def _fixture() -> dict[str, object]:
    return json.loads(FIXTURE.read_text(encoding="utf-8"))


def _unit(payload: dict[str, object], unit_id: str) -> dict[str, object]:
    return next(item for item in payload["units"] if item["unit_id"] == unit_id)  # type: ignore[index]


def _locator(payload: dict[str, object], locator_id: str) -> dict[str, object]:
    return next(  # type: ignore[return-value]
        item for item in payload["locators"] if item["locator_id"] == locator_id  # type: ignore[index]
    )


def test_positive_structure_map_has_contiguous_pages_and_closed_references() -> None:
    payload = _fixture()

    validate_structure_map(payload)

    assert payload["page_count"] == 4
    assert len(payload["pages"]) == 4  # type: ignore[arg-type]
    table = _unit(payload, "unit-synthetic-table-ae-specification")
    assert table["locator_ids"] == [
        "loc-synthetic-p003-ae-table",
        "loc-synthetic-p004-ae-table",
    ]


def test_synthetic_pdf_and_xlsx_cover_p2b_structure_shapes(tmp_path: Path) -> None:
    import fitz
    from openpyxl import load_workbook

    pdf_path, xlsx_path = create_structure_fixtures(tmp_path)
    document = fitz.open(pdf_path)
    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        assert document.page_count == 4
        assert document.get_toc(simple=True) == [
            [1, "1 Synthetic Fundamentals", 2],
            [2, "1.1 Synthetic Adverse Events", 3],
        ]
        assert "AE Specification - Segment 1" in document[2].get_text()
        assert "AE Specification - Segment 2" in document[3].get_text()
        assert workbook.sheetnames == ["Datasets", "Variables"]
        assert list(workbook["Datasets"].values)[1][0] == "AE"
        assert [row[1] for row in list(workbook["Variables"].values)[1:]] == [
            "AETERM",
            "AEDECOD",
        ]
    finally:
        workbook.close()
        document.close()


def test_stable_ids_only_use_namespace_and_source_anchor() -> None:
    assert stable_structure_map_id("synthetic") == "structure-map-synthetic"
    assert stable_page_id("synthetic", 3) == "page-synthetic-p0003"
    assert stable_unit_id("synthetic", "domain-ae") == "unit-synthetic-domain-ae"
    assert stable_locator_id("synthetic", "p003-domain-ae") == (
        "loc-synthetic-p003-domain-ae"
    )
    assert stable_reference_id("synthetic", "aeterm-to-ae") == (
        "ref-synthetic-aeterm-to-ae"
    )

    with pytest.raises(ValueError, match="not canonical"):
        stable_unit_id("synthetic", "bbox=72,60,360,90")


def test_gold_set_locator_ids_can_be_preserved_by_source_anchor_profile() -> None:
    gold_set = json.loads(
        (
            ROOT / "tests" / "fixtures" / "knowledge" / "sdtmig34-gold-set.json"
        ).read_text(encoding="utf-8")
    )

    for unit in gold_set["units"]:
        locator_id = unit["locator"]["locator_id"]
        identity_key = locator_id.removeprefix("loc-sdtmig34-")
        assert stable_locator_id("sdtmig34", identity_key) == locator_id


def test_assumption_is_a_paragraph_role_not_a_new_p1_source_unit_type() -> None:
    schema = json.loads(
        (
            ROOT / "schemas" / "extraction" / "source-structure-map.schema.json"
        ).read_text(encoding="utf-8")
    )
    unit_properties = schema["$defs"]["StructureUnit"]["properties"]

    assert "assumption" not in unit_properties["unit_type"]["enum"]
    assert "assumption" in unit_properties["content_role"]["enum"]


def test_canonical_bytes_ignore_mapping_key_order_but_not_array_order() -> None:
    payload = _fixture()
    reversed_root = {key: payload[key] for key in reversed(payload)}

    assert canonical_structure_map_bytes(payload) == canonical_structure_map_bytes(
        reversed_root
    )
    assert structure_map_sha256(payload) == structure_map_sha256(reversed_root)


def test_structure_map_forbids_generation_timestamp_in_canonical_payload() -> None:
    payload = _fixture()
    payload["generated_at"] = "2026-07-15T00:00:00Z"

    with pytest.raises(StructureMapContractError, match="schema validation failed"):
        validate_structure_map(payload)


def test_deep_locator_projects_to_existing_p1_locator_contract() -> None:
    locator = _locator(_fixture(), "loc-synthetic-p003-domain-ae")
    projected = project_locator_to_p1(locator)
    assert set(projected) == set(P1_LOCATOR_FIELDS)

    schema = json.loads(
        (ROOT / "schemas" / "extraction" / "knowledge-extraction.schema.json").read_text(
            encoding="utf-8"
        )
    )
    locator_schema = deepcopy(schema["$defs"]["Locator"])
    locator_schema["$schema"] = schema["$schema"]
    locator_schema["$defs"] = schema["$defs"]
    Draft202012Validator(locator_schema).validate(projected)


def test_page_assignments_fail_on_gap_or_reordering() -> None:
    payload = _fixture()
    payload["pages"][2]["physical_page"] = 4  # type: ignore[index]

    with pytest.raises(StructureMapContractError, match="unique, contiguous"):
        validate_structure_map(payload)


def test_nonblank_page_requires_primary_unit_or_deferred_explanation() -> None:
    payload = _fixture()
    payload["pages"][1]["primary_unit_id"] = None  # type: ignore[index]

    with pytest.raises(StructureMapContractError, match="lacks a primary unit"):
        validate_structure_map(payload)


def test_unit_parent_cycle_fails_closed() -> None:
    payload = _fixture()
    _unit(payload, "unit-synthetic-chapter-1")["parent_unit_id"] = (
        "unit-synthetic-domain-ae"
    )

    with pytest.raises(StructureMapContractError, match="parent cycle"):
        validate_structure_map(payload)


def test_duplicate_source_order_within_artifact_fails_closed() -> None:
    payload = _fixture()
    _unit(payload, "unit-synthetic-domain-ae")["source_order"] = 1

    with pytest.raises(StructureMapContractError, match="duplicate source_order"):
        validate_structure_map(payload)


def test_locator_artifact_media_type_must_match_locator_type() -> None:
    payload = _fixture()
    locator = _locator(payload, "loc-synthetic-xlsx-variables-ae-aeterm")
    locator.update(
        {
            "locator_type": "pdf_region",
            "physical_page": 3,
            "printed_page": "2",
            "bbox": [72.0, 60.0, 360.0, 90.0],
            "sheet_name": None,
            "row_number": None,
            "row_key": None,
        }
    )

    with pytest.raises(StructureMapContractError, match="pdf locator must reference"):
        validate_structure_map(payload)


@pytest.mark.parametrize(
    "bbox",
    [
        [360.0, 60.0, 72.0, 90.0],
        [72.0, 60.0, 700.0, 90.0],
    ],
)
def test_pdf_bbox_must_be_ordered_and_inside_page(bbox: list[float]) -> None:
    payload = _fixture()
    _locator(payload, "loc-synthetic-p003-domain-ae")["bbox"] = bbox

    with pytest.raises(StructureMapContractError, match="bbox"):
        validate_structure_map(payload)


def test_cross_page_table_segments_must_be_contiguous_and_ordered() -> None:
    payload = _fixture()
    _locator(payload, "loc-synthetic-p004-ae-table")["physical_page"] = 2

    with pytest.raises(StructureMapContractError, match="not contiguous"):
        validate_structure_map(payload)


def test_locator_table_id_must_resolve_to_table_unit() -> None:
    payload = _fixture()
    _locator(payload, "loc-synthetic-p003-ae-table")["table_id"] = (
        "unit-synthetic-domain-ae"
    )

    with pytest.raises(StructureMapContractError, match="dangling table_id"):
        validate_structure_map(payload)


def test_deferred_unit_requires_processing_note() -> None:
    payload = _fixture()
    unit = _unit(payload, "unit-synthetic-domain-ae")
    unit["processing_status"] = "deferred"
    unit["processing_note"] = None

    with pytest.raises(StructureMapContractError, match="processing_note"):
        validate_structure_map(payload)


def test_resolved_reference_target_must_exist() -> None:
    payload = _fixture()
    payload["references"][0]["to_unit_id"] = "unit-synthetic-missing"  # type: ignore[index]

    with pytest.raises(StructureMapContractError, match="dangling target"):
        validate_structure_map(payload)


def test_unresolved_reference_requires_reason_and_no_target() -> None:
    payload = _fixture()
    reference = payload["references"][0]  # type: ignore[index]
    reference["resolution_status"] = "unresolved"
    reference["target_kind"] = "unresolved"
    reference["to_unit_id"] = None
    reference["processing_note"] = None

    with pytest.raises(StructureMapContractError, match="processing_note"):
        validate_structure_map(payload)
